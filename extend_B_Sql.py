import mysql.connector
import time
from openpyxl import Workbook
from connection.connect_to_mysql import connect_to_mysql
from collections import defaultdict
from find_union_B import find_union  # Importing union function
import argparse

class Timer(object):
    def __init__(self, verbose=False):
        self.verbose = verbose

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, *args):
        self.end = time.perf_counter()
        self.secs = self.end - self.start
        self.msecs = self.secs * 1000  # millisecs
        if self.verbose:
            print('elapsed time: %f ms' % self.msecs)

def get_table_columns(cursor, table_name):
    try:
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = [column[0] for column in cursor.fetchall()]
        return columns
    except Exception as e:
        print(f"Error fetching columns from {table_name}: {e}")
        return []

def read_intervals_from_mysql(cursor, table_name, start_col, end_col):
    try:
        query = f"SELECT {start_col}, {end_col} FROM {table_name}"
        cursor.execute(query)
        intervals = cursor.fetchall()
        return intervals  # Assuming intervals are stored as (start, end) pairs
    except Exception as e:
        print(f"Error reading intervals from {table_name}: {e}")
        return []

def main():
    try:
        # Prompt the user for MySQL connection details
        host = input("Enter MySQL host address: ")
        user = input("Enter MySQL username: ")
        password = input("Enter MySQL password: ")
        database = input("Enter MySQL database name: ")
        table_name = input("Enter name of the table to process: ")

        # Establish a connection to MySQL using user-provided details
        db_connection = connect_to_mysql(
            host=host,
            user=user,
            password=password,
            database=database
        )
        cursor = db_connection.cursor()

        with Timer() as timer:
            # Create a dictionary to store union intervals for each pair of columns
            union_intervals_dict = defaultdict(list)
            
            # Filter columns ending with "_start" and "_end"
            start_col_A = "col1_start"
            end_col_A = "col1_end"
            
            for i in range(2, 5):
                start_col_B = f"col{i}_start"
                end_col_B = f"col{i}_end"
                
                # Read intervals from MySQL
                intervals_A = read_intervals_from_mysql(cursor, table_name, start_col_A, end_col_A)
                intervals_B = read_intervals_from_mysql(cursor, table_name, start_col_B, end_col_B)

                # Find union intervals between intervals_A and intervals_B
                union_intervals = find_union(intervals_A, intervals_B)
                
                if union_intervals:
                    # Print union intervals in the terminal
                    print(f"Union intervals between A and {i}:")
                    for interval_pair in union_intervals:
                        print(f"A{interval_pair[0]} > {i}{interval_pair[1]}")
                    
                    # Store union intervals in the dictionary
                    union_intervals_dict[f"{start_col_A} > {start_col_B}"].extend(union_intervals)

        cursor.close()
        db_connection.close()
        
        # Exporting to Excel
        wb = Workbook()
        ws = wb.active
        
        # Write union intervals to Excel
        for col_index, (column_name, intervals) in enumerate(union_intervals_dict.items(), start=1):
            ws.cell(row=1, column=col_index, value=column_name)
            row_index = 2
            for interval_pair in intervals:
                ws.cell(row=row_index, column=col_index, value=f"{interval_pair[0]} {interval_pair[1]}")
                row_index += 1
        
        wb.save("extend_A_Mysql.xlsx")

        print(f"Total running time: {timer.secs:.6f} seconds")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
