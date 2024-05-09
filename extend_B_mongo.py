import time
import csv
from connection.mongo_connection import connect_to_mongodb
from openpyxl import Workbook
from collections import defaultdict
from find_union_B import find_union  # Importing union function


class Timer(object):
    def __init__(self, verbose=False):
        self.verbose = verbose

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, *args):
        self.end = time.perf_counter()
        self.secs = self.end - self.start
        self.msecs = self.secs * 1000  # millisecs
        if self.verbose:
            print('elapsed time: %f ms' % self.msecs)



def print_intervals_without_gaps(intervals):
    sorted_intervals = sorted(intervals, key=lambda x: (x[0], x[1]))
    current_line = []
    seen_intervals = set()  # To track seen intervals
    output_lines = []  # Store output lines
    for interval in sorted_intervals:
        # Check if interval is a tuple before accessing its elements
        if isinstance(interval, tuple) and len(interval) >= 2:
            if interval not in seen_intervals:  # Check if interval is already seen
                if not current_line or interval[0] <= current_line[-1][1]:
                    current_line.append(interval)
                else:
                    output_lines.append(current_line)
                    current_line = [interval]
                seen_intervals.add(interval)  # Add interval to seen_intervals
        else:
            print("Invalid interval:", interval)
    if current_line:  # Print the remaining intervals if any
        output_lines.append(current_line)

    # Convert the intervals to strings before returning
    return [
        [f"({interval[0]}, {interval[1]})" for interval in line]
        for line in output_lines
    ]


def print_line_with_commas(intervals):
    for i, interval in enumerate(intervals):
        print(f"({interval[0]}, {interval[1]})", end=' ')
    print()


def get_table_columns(collection):
    try:
        # Get the first document in the collection
        sample_document = collection.find_one()

        # Extract table columns
        columns = list(sample_document.keys())
        
        # Remove the '_id' field if present
        if '_id' in columns:
            columns.remove('_id')

        return columns
    except Exception as e:
        print(f"Error getting table columns: {e}")
        return []

def print_line(intervals):
    for interval in intervals:
        print(f"{interval[0]} {interval[1]}", end=' ')
    print() 

def read_intervals_from_mongodb(collection, table_name):
    try:
        result = collection.find_one({}, {table_name: 1, '_id': 0})
        return result[table_name]
    except Exception as e:
        print(f"Error reading from MongoDB: {e}")
        return []


def select_tables_to_compare(tables):
    # Specify the criteria to select tables for comparison
    # For example, you might want to compare Table 1 with tables having indexes greater than 1
    return tables[1:]  # Here, we select all tables except Table 1

def main():
    # Prompt the user to enter the database and collection names
    database_name = input("Enter the name of the MongoDB database: ")
    collection_name = input("Enter the name of the MongoDB collection: ")

    # Connect to MongoDB
    database, collection = connect_to_mongodb(database_name, collection_name)
    
    output_lines = []  # Store output lines for exporting to CSV
    union_results = []  # Store all union intervals found

    with Timer() as timer:
        union_intervals_dict = defaultdict(list)
        tables = get_table_columns(collection)
        table_1 = tables[0]  # Assuming Table 1 is the first table in your collection

        # Specify the tables you want to compare Table 1 with based on some criteria
        tables_to_compare = select_tables_to_compare(tables)

        for table_B in tables_to_compare:
            # Read intervals from MongoDB for both tables
            intervals_A = read_intervals_from_mongodb(collection, table_1)
            intervals_B = read_intervals_from_mongodb(collection, table_B)

            # Find union intervals between intervals_A and intervals_B
            union_intervals = find_union(intervals_A, intervals_B)
            
            if union_intervals:
                print(f"Union intervals between {table_1} and {table_B}:")
                for interval_pair in union_intervals:
                    print(f"{table_1}{interval_pair[0]} > {table_B}{interval_pair[1]}")
                
                # Append to union_results
                union_results.extend(union_intervals)
                union_intervals_dict[f"{table_1} < {table_B}"].extend(union_intervals)

    # Exporting to Excel
    wb = Workbook()
    ws = wb.active
    
    # Write union intervals to Excel
    for col_index, (column_name, intervals) in enumerate(union_intervals_dict.items(), start=1):
        ws.cell(row=1, column=col_index, value=column_name)
        row_index = 2
        for interval_pair in intervals:
            ws.cell(row=row_index, column=col_index, value=f"{interval_pair[0]} {interval_pair[1]}")
            row_index += 1
    
    wb.save("extend_A_Nosql.xlsx")

    print(f"Total running time: {timer.secs:.6f} seconds")


if __name__ == "__main__":
    main()

